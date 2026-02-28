import logging
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackQueryHandler
from datetime import datetime
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ========== ĞĞĞ¡Ğ¢Ğ ĞĞ™ĞšĞ˜ ==========
TOKEN = "8761306495:AAFWICUB62qgO2h-1va3Y50DHZPGvCGakjw"
DATA_FILE = "gabbana_data.json"
EXCEL_FILE = "gabbana_budget.xlsx"
ALLOWED_IDS = [6578266978, 5029738209, 7950080109]

# Ğ¡Ğ¿Ğ¸ÑĞ¾Ğº ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ¾Ğ²
EMPLOYEES = ["ĞœĞ°Ñ‚Ğ²ĞµĞ¹", "Ğ”Ğ¸Ğ¼Ğ°", "ĞĞ¸ĞºĞ¸Ñ‚Ğ°"]

# Ğ›Ğ¾Ğ³Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ğµ
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ñ‹
main_keyboard = ReplyKeyboardMarkup([
    ['ğŸ’° Ğ”Ğ¾Ñ…Ğ¾Ğ´', 'ğŸ’¸ Ğ Ğ°ÑÑ…Ğ¾Ğ´'],
    ['ğŸ“Š Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ°', 'ğŸ“‹ Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ¾Ğ²'],
    ['ğŸ‘¥ Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ° ĞºĞ¾Ğ»Ğ»ĞµĞ³', 'âœï¸ Ğ ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ñ‚ÑŒ/Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ']
], resize_keyboard=True)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ğ° Ñ ĞºĞ½Ğ¾Ğ¿ĞºĞ¾Ğ¹ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹
cancel_keyboard = ReplyKeyboardMarkup([['ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°']], resize_keyboard=True)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ğ° Ğ´Ğ»Ñ Ğ¾Ğ±ÑŠĞµĞ¼Ğ°
volume_keyboard = ReplyKeyboardMarkup([
    ['6ml', '10ml'],
    ['ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°']
], resize_keyboard=True)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ğ° Ğ´Ğ»Ñ ÑĞ¿Ğ¾ÑĞ¾Ğ±Ğ° Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹
payment_keyboard = ReplyKeyboardMarkup([
    ['ğŸ’³ ĞŸĞµÑ€ĞµĞ²Ğ¾Ğ´', 'ğŸ’µ ĞĞ°Ğ»Ğ¸Ñ‡ĞºĞ°'],
    ['ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°']
], resize_keyboard=True)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ğ° Ğ´Ğ»Ñ Ğ±Ğ°Ğ½ĞºĞ¾Ğ²
bank_keyboard = ReplyKeyboardMarkup([
    ['ğŸ¦ Ğ¡Ğ±ĞµÑ€', 'ğŸ¦ Ğ¢Ğ¸Ğ½ÑŒĞºĞ¾Ñ„Ñ„', 'ğŸ¦ Ğ’Ğ¢Ğ‘'],
    ['ğŸ¦ ĞĞ»ÑŒÑ„Ğ°', 'ğŸ¦ Ğ Ğ°Ğ¹Ñ„Ñ„Ğ°Ğ¹Ğ·ĞµĞ½', 'ğŸ¦ Ğ”Ñ€ÑƒĞ³Ğ¾Ğ¹'],
    ['ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°']
], resize_keyboard=True)

# ĞšĞ»Ğ°Ğ²Ğ¸Ğ°Ñ‚ÑƒÑ€Ğ° Ğ´Ğ»Ñ Ğ²Ñ‹Ğ±Ğ¾Ñ€Ğ° ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°
employee_keyboard = ReplyKeyboardMarkup([
    ['ğŸ‘¤ ĞœĞ°Ñ‚Ğ²ĞµĞ¹', 'ğŸ‘¤ Ğ”Ğ¸Ğ¼Ğ°', 'ğŸ‘¤ ĞĞ¸ĞºĞ¸Ñ‚Ğ°'],
    ['ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°']
], resize_keyboard=True)

# Ğ¡Ğ¾ÑÑ‚Ğ¾ÑĞ½Ğ¸Ñ Ğ´Ğ»Ñ Ğ´Ğ¾Ñ…Ğ¾Ğ´Ğ°
INCOME_STATES = {
    'NAME': 1,
    'VOLUME': 2,
    'QUANTITY': 3,
    'EMPLOYEE': 4,
    'PAYMENT': 5,
    'BANK': 6,
    'AMOUNT': 7
}

# Ğ¡Ğ¾ÑÑ‚Ğ¾ÑĞ½Ğ¸Ñ Ğ´Ğ»Ñ Ñ€Ğ°ÑÑ…Ğ¾Ğ´Ğ°
EXPENSE_STATES = {
    'AMOUNT': 1,
    'DESCRIPTION': 2,
    'EMPLOYEE': 3
}

# Ğ¡Ğ¾ÑÑ‚Ğ¾ÑĞ½Ğ¸Ñ Ğ¿Ğ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ĞµĞ¹
user_data = {}

def check_access(update):
    user_id = update.effective_user.id
    return user_id in ALLOWED_IDS

def load_data():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {'operations': [], 'next_id': 1}
    return {'operations': [], 'next_id': 1}

def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_next_id():
    data = load_data()
    next_id = data.get('next_id', 1)
    data['next_id'] = next_id + 1
    save_data(data)
    return next_id

def add_operation(operation):
    data = load_data()
    if 'operations' not in data:
        data['operations'] = []
    data['operations'].append(operation)
    save_data(data)
    save_to_excel(operation)

def get_all_operations():
    data = load_data()
    return data.get('operations', [])

def delete_operation(op_id):
    data = load_data()
    data['operations'] = [op for op in data['operations'] if op['id'] != op_id]
    save_data(data)

def update_operation(op_id, updated_op):
    data = load_data()
    for i, op in enumerate(data['operations']):
        if op['id'] == op_id:
            data['operations'][i] = updated_op
            break
    save_data(data)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Gabbana&Home'
        
        headers = ['ID', 'Ğ”Ğ°Ñ‚Ğ°', 'Ğ¢Ğ¸Ğ¿', 'ĞŸĞ°Ñ€Ñ„ÑĞ¼', 'ĞĞ±ÑŠĞµĞ¼', 'ĞšĞ¾Ğ»-Ğ²Ğ¾', 'Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº', 'Ğ¡Ğ¿Ğ¾ÑĞ¾Ğ± Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹', 'Ğ‘Ğ°Ğ½Ğº', 'Ğ¡ÑƒĞ¼Ğ¼Ğ° (â‚½)', 'ĞĞ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ', 'ĞšÑ‚Ğ¾ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ¸Ğ»']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 35
        ws.column_dimensions['L'].width = 15
        
        wb.save(EXCEL_FILE)

def save_to_excel(operation):
    try:
        if not os.path.exists(EXCEL_FILE):
            init_excel()
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        last_row = ws.max_row + 1
        
        ws.cell(row=last_row, column=1, value=operation['id'])
        ws.cell(row=last_row, column=2, value=operation['date'])
        ws.cell(row=last_row, column=3, value=operation['type_display'])
        ws.cell(row=last_row, column=4, value=operation.get('parfum_name', '-'))
        ws.cell(row=last_row, column=5, value=operation.get('volume', '-'))
        ws.cell(row=last_row, column=6, value=operation.get('quantity', 0))
        ws.cell(row=last_row, column=7, value=operation.get('employee', '-'))
        ws.cell(row=last_row, column=8, value=operation.get('payment', '-'))
        ws.cell(row=last_row, column=9, value=operation.get('bank', '-'))
        ws.cell(row=last_row, column=10, value=operation['amount'])
        ws.cell(row=last_row, column=11, value=operation.get('description', ''))
        ws.cell(row=last_row, column=12, value=operation.get('added_by', ''))
        
        ws.cell(row=last_row, column=10).number_format = '#,##0.00 â‚½'
        
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"ĞÑˆĞ¸Ğ±ĞºĞ° ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ğ¸Ñ Ğ² Excel: {e}")

def start(update, context):
    if not check_access(update):
        update.message.reply_text("âŒ Ğ£ Ğ²Ğ°Ñ Ğ½ĞµÑ‚ Ğ´Ğ¾ÑÑ‚ÑƒĞ¿Ğ° Ğº ÑÑ‚Ğ¾Ğ¼Ñƒ Ğ±Ğ¾Ñ‚Ñƒ")
        return
    
    user = update.effective_user
    
    welcome_text = (
        f"âœ¨ *Gabbana&Home Parfum* âœ¨\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ‘‹ *Ğ”Ğ¾Ğ±Ñ€Ğ¾ Ğ¿Ğ¾Ğ¶Ğ°Ğ»Ğ¾Ğ²Ğ°Ñ‚ÑŒ, {user.first_name}!*\n\n"
        f"ğŸ“Š *ĞŸĞ°Ñ€Ñ„ÑĞ¼ĞµÑ€Ğ½Ñ‹Ğ¹ ÑƒÑ‡ĞµÑ‚*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"âœ… *ĞÑĞ½Ğ¾Ğ²Ğ½Ñ‹Ğµ Ñ„ÑƒĞ½ĞºÑ†Ğ¸Ğ¸:*\n"
        f"â€¢ ğŸ’° Ğ”Ğ¾Ñ…Ğ¾Ğ´ - Ğ¿Ñ€Ğ¾Ğ´Ğ°Ğ¶Ğ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ°\n"
        f"â€¢ ğŸ’¸ Ğ Ğ°ÑÑ…Ğ¾Ğ´ - Ğ·Ğ°ĞºÑƒĞ¿ĞºĞ°/Ñ€Ğ°ÑÑ…Ğ¾Ğ´Ñ‹\n"
        f"â€¢ ğŸ“Š Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ° - Ğ¾Ğ±Ñ‰Ğ°Ñ ÑÑ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ°\n"
        f"â€¢ ğŸ“‹ Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ¾Ğ² - Ğ²ÑĞµ Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ñ‹\n"
        f"â€¢ ğŸ‘¥ Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ° ĞºĞ¾Ğ»Ğ»ĞµĞ³ - Ğ¿Ğ¾ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°Ğ¼\n"
        f"â€¢ âœï¸ Ğ ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ñ‚ÑŒ/Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ\n\n"
        f"âœ¨ *Ğ’ÑĞµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ÑÑÑ‚ÑÑ*"
    )
    
    update.message.reply_text(welcome_text, parse_mode='Markdown', reply_markup=main_keyboard)

def handle_income(update, context):
    if not check_access(update):
        return
    
    chat_id = update.effective_chat.id
    
    user_data[chat_id] = {
        'type': 'income',
        'state': INCOME_STATES['NAME'],
        'added_by': update.effective_user.first_name
    }
    
    update.message.reply_text(
        "ğŸ’µ *Ğ”ĞĞ¥ĞĞ” (ĞŸÑ€Ğ¾Ğ´Ğ°Ğ¶Ğ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ°)*\n"
        "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        "ğŸ“ *Ğ¨Ğ°Ğ³ 1 Ğ¸Ğ· 7*\n\n"
        "âœï¸ *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ğ½Ğ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ°:*\n\n"
        "ğŸ’¡ *ĞŸÑ€Ğ¸Ğ¼ĞµÑ€Ñ‹:*\n"
        "â€¢ Creed Aventus\n"
        "â€¢ Baccarat Rouge 540\n"
        "â€¢ Tom Ford Tobacco Vanille\n\n"
        "ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
        parse_mode='Markdown',
        reply_markup=cancel_keyboard
    )

def handle_income_name(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    user_data[chat_id]['parfum_name'] = text
    user_data[chat_id]['state'] = INCOME_STATES['VOLUME']
    
    update.message.reply_text(
        f"âœ… ĞĞ°Ğ·Ğ²Ğ°Ğ½Ğ¸Ğµ: *{text}*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ“ *Ğ¨Ğ°Ğ³ 2 Ğ¸Ğ· 7*\n\n"
        f"ğŸ”¢ *Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ Ğ¾Ğ±ÑŠĞµĞ¼:*\n\n"
        f"ğŸ’¡ *Ğ”Ğ¾ÑÑ‚ÑƒĞ¿Ğ½Ñ‹Ğµ Ğ¾Ğ±ÑŠĞµĞ¼Ñ‹:* 6ml, 10ml",
        parse_mode='Markdown',
        reply_markup=volume_keyboard
    )

def handle_income_volume(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    if text not in ['6ml', '10ml']:
        update.message.reply_text(
            "âŒ ĞŸĞ¾Ğ¶Ğ°Ğ»ÑƒĞ¹ÑÑ‚Ğ°, Ğ²Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ Ğ¾Ğ±ÑŠĞµĞ¼ Ğ¸Ğ· ĞºĞ½Ğ¾Ğ¿Ğ¾Ğº:",
            reply_markup=volume_keyboard
        )
        return
    
    user_data[chat_id]['volume'] = text
    user_data[chat_id]['state'] = INCOME_STATES['QUANTITY']
    
    update.message.reply_text(
        f"âœ… ĞĞ±ÑŠĞµĞ¼: *{text}*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ“ *Ğ¨Ğ°Ğ³ 3 Ğ¸Ğ· 7*\n\n"
        f"ğŸ”¢ *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ĞºĞ¾Ğ»Ğ¸Ñ‡ĞµÑÑ‚Ğ²Ğ¾ Ñ„Ğ»Ğ°ĞºĞ¾Ğ½Ğ¾Ğ²:*\n\n"
        f"ğŸ’¡ *ĞŸÑ€Ğ¸Ğ¼ĞµÑ€Ñ‹:* 1, 2, 3\n\n"
        f"ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
        parse_mode='Markdown',
        reply_markup=cancel_keyboard
    )

def handle_income_quantity(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    try:
        quantity = int(text)
        if quantity <= 0:
            raise ValueError
        
        user_data[chat_id]['quantity'] = quantity
        user_data[chat_id]['state'] = INCOME_STATES['EMPLOYEE']
        
        update.message.reply_text(
            f"âœ… ĞšĞ¾Ğ»Ğ¸Ñ‡ĞµÑÑ‚Ğ²Ğ¾: *{quantity} ÑˆÑ‚*\n"
            f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
            f"ğŸ“ *Ğ¨Ğ°Ğ³ 4 Ğ¸Ğ· 7*\n\n"
            f"ğŸ‘¤ *Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°:*",
            parse_mode='Markdown',
            reply_markup=employee_keyboard
        )
        
    except ValueError:
        update.message.reply_text(
            "âŒ Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½Ğ¾Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾\n\nğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
            parse_mode='Markdown',
            reply_markup=cancel_keyboard
        )

def handle_income_employee(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    employee = text.replace('ğŸ‘¤ ', '')
    
    if employee not in EMPLOYEES:
        update.message.reply_text(
            "âŒ ĞŸĞ¾Ğ¶Ğ°Ğ»ÑƒĞ¹ÑÑ‚Ğ°, Ğ²Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ° Ğ¸Ğ· ĞºĞ½Ğ¾Ğ¿Ğ¾Ğº:",
            reply_markup=employee_keyboard
        )
        return
    
    user_data[chat_id]['employee'] = employee
    user_data[chat_id]['state'] = INCOME_STATES['PAYMENT']
    
    update.message.reply_text(
        f"âœ… Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº: *{employee}*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ“ *Ğ¨Ğ°Ğ³ 5 Ğ¸Ğ· 7*\n\n"
        f"ğŸ’³ *Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¿Ğ¾ÑĞ¾Ğ± Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹:*",
        parse_mode='Markdown',
        reply_markup=payment_keyboard
    )

def handle_income_payment(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    if 'ĞŸĞµÑ€ĞµĞ²Ğ¾Ğ´' in text:
        user_data[chat_id]['payment'] = 'ĞŸĞµÑ€ĞµĞ²Ğ¾Ğ´'
        user_data[chat_id]['state'] = INCOME_STATES['BANK']
        
        update.message.reply_text(
            f"âœ… Ğ¡Ğ¿Ğ¾ÑĞ¾Ğ± Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹: *{text}*\n"
            f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
            f"ğŸ“ *Ğ¨Ğ°Ğ³ 6 Ğ¸Ğ· 7*\n\n"
            f"ğŸ¦ *Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ Ğ±Ğ°Ğ½Ğº:*",
            parse_mode='Markdown',
            reply_markup=bank_keyboard
        )
        
    elif 'ĞĞ°Ğ»Ğ¸Ñ‡ĞºĞ°' in text:
        user_data[chat_id]['payment'] = 'ĞĞ°Ğ»Ğ¸Ñ‡ĞºĞ°'
        user_data[chat_id]['bank'] = '-'
        user_data[chat_id]['state'] = INCOME_STATES['AMOUNT']
        
        update.message.reply_text(
            f"âœ… Ğ¡Ğ¿Ğ¾ÑĞ¾Ğ± Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹: *{text}*\n"
            f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
            f"ğŸ“ *Ğ¨Ğ°Ğ³ 7 Ğ¸Ğ· 7*\n\n"
            f"ğŸ’° *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ÑÑƒĞ¼Ğ¼Ñƒ Ğ´Ğ¾Ñ…Ğ¾Ğ´Ğ°:*\n\n"
            f"ğŸ“ *Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ñ‹:* 1300, 2 500, 3 000.50\n\n"
            f"ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
            parse_mode='Markdown',
            reply_markup=cancel_keyboard
        )
    else:
        update.message.reply_text(
            "âŒ ĞŸĞ¾Ğ¶Ğ°Ğ»ÑƒĞ¹ÑÑ‚Ğ°, Ğ²Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¿Ğ¾ÑĞ¾Ğ± Ğ¾Ğ¿Ğ»Ğ°Ñ‚Ñ‹ Ğ¸Ğ· ĞºĞ½Ğ¾Ğ¿Ğ¾Ğº:",
            reply_markup=payment_keyboard
        )

def handle_income_bank(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    bank = text.replace('ğŸ¦ ', '') if 'ğŸ¦' in text else text
    
    user_data[chat_id]['bank'] = bank
    user_data[chat_id]['state'] = INCOME_STATES['AMOUNT']
    
    update.message.reply_text(
        f"âœ… Ğ‘Ğ°Ğ½Ğº: *{bank}*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ“ *Ğ¨Ğ°Ğ³ 7 Ğ¸Ğ· 7*\n\n"
        f"ğŸ’° *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ÑÑƒĞ¼Ğ¼Ñƒ Ğ´Ğ¾Ñ…Ğ¾Ğ´Ğ°:*\n\n"
        f"ğŸ“ *Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ñ‹:* 1300, 2 500, 3 000.50\n\n"
        f"ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
        parse_mode='Markdown',
        reply_markup=cancel_keyboard
    )

def handle_income_amount(update, context):
    chat_id = update.effective_chat.id
    user = update.effective_user
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    try:
        amount = float(text.replace(' ', '').replace(',', '.'))
        
        data = user_data.pop(chat_id)
        
        operation = {
            'id': get_next_id(),
            'date': datetime.now().strftime('%d.%m.%Y %H:%M'),
            'type': 'income',
            'type_display': 'ğŸ’° Ğ”Ğ¾Ñ…Ğ¾Ğ´',
            'parfum_name': data['parfum_name'],
            'volume': data['volume'],
            'quantity': data['quantity'],
            'employee': data['employee'],
            'payment': data['payment'],
            'bank': data.get('bank', '-'),
            'amount': amount,
            'description': f"{data['parfum_name']} {data['volume']} x{data['quantity']}",
            'added_by': data['added_by']
        }
        
        add_operation(operation)
        
        formatted_amount = f"{amount:,.0f} â‚½".replace(',', ' ')
        if amount != int(amount):
            formatted_amount = f"{amount:,.2f} â‚½".replace(',', ' ')
        
        report = (
            f"âœ… *ĞŸĞ ĞĞ”ĞĞ–Ğ #{operation['id']} Ğ—ĞĞŸĞ˜Ğ¡ĞĞĞ!*\n"
            f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
            f"ğŸ“Œ *ĞŸĞ°Ñ€Ñ„ÑĞ¼:* {data['parfum_name']}\n"
            f"ğŸ“Œ *ĞĞ±ÑŠĞµĞ¼:* {data['volume']}\n"
            f"ğŸ“Œ *ĞšĞ¾Ğ»-Ğ²Ğ¾:* {data['quantity']} ÑˆÑ‚\n"
            f"ğŸ‘¤ *Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº:* {data['employee']}\n"
            f"ğŸ’³ *ĞĞ¿Ğ»Ğ°Ñ‚Ğ°:* {data['payment']}\n"
        )
        
        if data['payment'] == 'ĞŸĞµÑ€ĞµĞ²Ğ¾Ğ´':
            report += f"ğŸ¦ *Ğ‘Ğ°Ğ½Ğº:* {data.get('bank', '-')}\n"
        
        report += f"ğŸ’° *Ğ¡ÑƒĞ¼Ğ¼Ğ°:* {formatted_amount}\n"
        report += f"ğŸ“ *Ğ—Ğ°Ğ¿Ğ¸ÑÑŒ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ¸Ğ»:* {data['added_by']}\n\n"
        report += f"âœ¨ *Ğ¡Ğ¿Ğ°ÑĞ¸Ğ±Ğ¾ Ğ·Ğ° Ğ¿Ñ€Ğ¾Ğ´Ğ°Ğ¶Ñƒ!*"
        
        update.message.reply_text(report, parse_mode='Markdown', reply_markup=main_keyboard)
        
    except ValueError:
        update.message.reply_text(
            "âŒ Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½ÑƒÑ ÑÑƒĞ¼Ğ¼Ñƒ\n\nğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
            parse_mode='Markdown',
            reply_markup=cancel_keyboard
        )

def handle_expense(update, context):
    if not check_access(update):
        return
    
    chat_id = update.effective_chat.id
    
    user_data[chat_id] = {
        'type': 'expense',
        'state': EXPENSE_STATES['AMOUNT'],
        'added_by': update.effective_user.first_name
    }
    
    update.message.reply_text(
        "ğŸ’³ *Ğ ĞĞ¡Ğ¥ĞĞ”*\n"
        "â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        "âœï¸ *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ÑÑƒĞ¼Ğ¼Ñƒ Ñ€Ğ°ÑÑ…Ğ¾Ğ´Ğ°:*\n\n"
        "ğŸ“ *Ğ¤Ğ¾Ñ€Ğ¼Ğ°Ñ‚Ñ‹:* 1300, 2 500, 3 000.50\n\n"
        "ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
        parse_mode='Markdown',
        reply_markup=cancel_keyboard
    )

def handle_expense_amount(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    try:
        amount = float(text.replace(' ', '').replace(',', '.'))
        user_data[chat_id]['amount'] = amount
        user_data[chat_id]['state'] = EXPENSE_STATES['DESCRIPTION']
        
        formatted_amount = f"{amount:,.0f} â‚½".replace(',', ' ')
        if amount != int(amount):
            formatted_amount = f"{amount:,.2f} â‚½".replace(',', ' ')
        
        update.message.reply_text(
            f"âœ… *Ğ¡ÑƒĞ¼Ğ¼Ğ°:* {formatted_amount}\n"
            f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
            f"ğŸ“ *Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ğ¾Ğ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ Ñ€Ğ°ÑÑ…Ğ¾Ğ´Ğ°:*\n\n"
            f"ğŸ’¡ ĞĞ°Ğ¿Ñ€Ğ¸Ğ¼ĞµÑ€: Ğ—Ğ°ĞºÑƒĞ¿ĞºĞ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ°, ĞÑ€ĞµĞ½Ğ´Ğ°, Ğ ĞµĞºĞ»Ğ°Ğ¼Ğ°\n\n"
            f"ğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
            parse_mode='Markdown',
            reply_markup=cancel_keyboard
        )
    except ValueError:
        update.message.reply_text(
            "âŒ Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ ĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½ÑƒÑ ÑÑƒĞ¼Ğ¼Ñƒ\n\nğŸ”¹ *Ğ”Ğ»Ñ Ğ¾Ñ‚Ğ¼ĞµĞ½Ñ‹ Ğ½Ğ°Ğ¶Ğ¼Ğ¸Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºÑƒ Ğ½Ğ¸Ğ¶Ğµ*",
            parse_mode='Markdown',
            reply_markup=cancel_keyboard
        )

def handle_expense_description(update, context):
    chat_id = update.effective_chat.id
    description = update.message.text
    
    if description == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    user_data[chat_id]['description'] = description
    user_data[chat_id]['state'] = EXPENSE_STATES['EMPLOYEE']
    
    update.message.reply_text(
        f"âœ… ĞĞ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ: *{description}*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ‘¤ *Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°:*",
        parse_mode='Markdown',
        reply_markup=employee_keyboard
    )

def handle_expense_employee(update, context):
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        del user_data[chat_id]
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    employee = text.replace('ğŸ‘¤ ', '')
    
    if employee not in EMPLOYEES:
        update.message.reply_text(
            "âŒ ĞŸĞ¾Ğ¶Ğ°Ğ»ÑƒĞ¹ÑÑ‚Ğ°, Ğ²Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ° Ğ¸Ğ· ĞºĞ½Ğ¾Ğ¿Ğ¾Ğº:",
            reply_markup=employee_keyboard
        )
        return
    
    data = user_data.pop(chat_id)
    
    operation = {
        'id': get_next_id(),
        'date': datetime.now().strftime('%d.%m.%Y %H:%M'),
        'type': 'expense',
        'type_display': 'ğŸ’¸ Ğ Ğ°ÑÑ…Ğ¾Ğ´',
        'amount': data['amount'],
        'description': data['description'],
        'employee': employee,
        'added_by': data['added_by']
    }
    
    add_operation(operation)
    
    formatted_amount = f"{data['amount']:,.0f} â‚½".replace(',', ' ')
    if data['amount'] != int(data['amount']):
        formatted_amount = f"{data['amount']:,.2f} â‚½".replace(',', ' ')
    
    update.message.reply_text(
        f"âœ… *Ğ ĞĞ¡Ğ¥ĞĞ” #{operation['id']} Ğ—ĞĞŸĞ˜Ğ¡ĞĞ!*\n"
        f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        f"ğŸ’° *Ğ¡ÑƒĞ¼Ğ¼Ğ°:* {formatted_amount}\n"
        f"ğŸ“‹ *ĞĞ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ:* {data['description']}\n"
        f"ğŸ‘¤ *Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº:* {employee}\n"
        f"ğŸ“ *Ğ—Ğ°Ğ¿Ğ¸ÑÑŒ Ğ´Ğ¾Ğ±Ğ°Ğ²Ğ¸Ğ»:* {data['added_by']}",
        parse_mode='Markdown',
        reply_markup=main_keyboard
    )

def show_parfum_table(update, context):
    if not check_access(update):
        return
    
    operations = get_all_operations()
    
    parfums = {}
    for op in operations:
        if op['type'] == 'income':
            key = f"{op['parfum_name']} ({op['volume']})"
            if key not in parfums:
                parfums[key] = {
                    'name': op['parfum_name'],
                    'volume': op['volume'],
                    'total_quantity': 0,
                    'total_amount': 0,
                    'sales': []
                }
            parfums[key]['total_quantity'] += op['quantity']
            parfums[key]['total_amount'] += op['amount']
            parfums[key]['sales'].append(op)
    
    if not parfums:
        update.message.reply_text("ğŸ“­ ĞĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ… Ğ¾ Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ°Ñ…", reply_markup=main_keyboard)
        return
    
    sorted_parfums = sorted(parfums.items(), key=lambda x: x[1]['total_amount'], reverse=True)
    
    report = "ğŸ“‹ *Ğ¢ĞĞ‘Ğ›Ğ˜Ğ¦Ğ ĞŸĞĞ Ğ¤Ğ®ĞœĞĞ’*\n"
    report += "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
    
    for key, data in sorted_parfums:
        formatted_amount = f"{data['total_amount']:,.0f} â‚½".replace(',', ' ')
        report += (
            f"ğŸ“Œ *{data['name']}*\n"
            f"   â€¢ ĞĞ±ÑŠĞµĞ¼: {data['volume']}\n"
            f"   â€¢ ĞŸÑ€Ğ¾Ğ´Ğ°Ğ½Ğ¾: {data['total_quantity']} ÑˆÑ‚\n"
            f"   â€¢ ĞĞ° ÑÑƒĞ¼Ğ¼Ñƒ: `{formatted_amount}`\n"
            f"   â€¢ ĞŸÑ€Ğ¾Ğ´Ğ°Ğ¶: {len(data['sales'])}\n\n"
        )
    
    ml6_total = sum(data['total_amount'] for key, data in parfums.items() if '6ml' in key)
    ml10_total = sum(data['total_amount'] for key, data in parfums.items() if '10ml' in key)
    
    report += "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n"
    report += f"ğŸ“Š *6ml:* {ml6_total:,.0f} â‚½\n".replace(',', ' ')
    report += f"ğŸ“Š *10ml:* {ml10_total:,.0f} â‚½\n".replace(',', ' ')
    
    update.message.reply_text(report, parse_mode='Markdown', reply_markup=main_keyboard)

def show_employee_stats(update, context):
    if not check_access(update):
        return
    
    operations = get_all_operations()
    
    if not operations:
        update.message.reply_text("ğŸ“­ ĞĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ…", reply_markup=main_keyboard)
        return
    
    stats = {}
    for employee in EMPLOYEES:
        stats[employee] = {
            'income': 0,
            'income_count': 0,
            'expense': 0,
            'expense_count': 0,
            'parfums': {}
        }
    
    for op in operations:
        employee = op.get('employee', 'ĞĞµ ÑƒĞºĞ°Ğ·Ğ°Ğ½')
        if employee in stats:
            if op['type'] == 'income':
                stats[employee]['income'] += op['amount']
                stats[employee]['income_count'] += 1
                
                key = f"{op['parfum_name']} {op['volume']}"
                if key not in stats[employee]['parfums']:
                    stats[employee]['parfums'][key] = {
                        'quantity': 0,
                        'amount': 0
                    }
                stats[employee]['parfums'][key]['quantity'] += op['quantity']
                stats[employee]['parfums'][key]['amount'] += op['amount']
                
            elif op['type'] == 'expense':
                stats[employee]['expense'] += op['amount']
                stats[employee]['expense_count'] += 1
    
    report = "ğŸ‘¥ *Ğ¡Ğ¢ĞĞ¢Ğ˜Ğ¡Ğ¢Ğ˜ĞšĞ ĞŸĞ Ğ¡ĞĞ¢Ğ Ğ£Ğ”ĞĞ˜ĞšĞĞœ*\n"
    report += "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
    
    for employee in EMPLOYEES:
        data = stats[employee]
        income_formatted = f"{data['income']:,.0f} â‚½".replace(',', ' ')
        expense_formatted = f"{data['expense']:,.0f} â‚½".replace(',', ' ')
        profit_formatted = f"{data['income'] - data['expense']:,.0f} â‚½".replace(',', ' ')
        
        report += f"ğŸ‘¤ *{employee}*\n"
        report += f"   ğŸ“ˆ Ğ”Ğ¾Ñ…Ğ¾Ğ´Ñ‹: `{income_formatted}` ({data['income_count']} ÑˆÑ‚)\n"
        report += f"   ğŸ“‰ Ğ Ğ°ÑÑ…Ğ¾Ğ´Ñ‹: `{expense_formatted}` ({data['expense_count']} ÑˆÑ‚)\n"
        report += f"   ğŸ’ Ğ˜Ñ‚Ğ¾Ğ³: `{profit_formatted}`\n"
        
        if data['parfums']:
            report += f"   ğŸ“¦ ĞŸÑ€Ğ¾Ğ´Ğ°Ğ¶Ğ¸:\n"
            top_parfums = sorted(data['parfums'].items(), key=lambda x: x[1]['amount'], reverse=True)[:3]
            for parfum, pdata in top_parfums:
                pamount = f"{pdata['amount']:,.0f} â‚½".replace(',', ' ')
                report += f"      â€¢ {parfum}: {pdata['quantity']} ÑˆÑ‚ ({pamount})\n"
        report += "\n"
    
    update.message.reply_text(report, parse_mode='Markdown', reply_markup=main_keyboard)

def show_all_statistics(update, context):
    if not check_access(update):
        return
    
    operations = get_all_operations()
    
    if not operations:
        update.message.reply_text("ğŸ“­ ĞĞµÑ‚ Ğ´Ğ°Ğ½Ğ½Ñ‹Ñ…", reply_markup=main_keyboard)
        return
    
    income_total = sum(op['amount'] for op in operations if op['type'] == 'income')
    expense_total = sum(op['amount'] for op in operations if op['type'] == 'expense')
    income_count = len([op for op in operations if op['type'] == 'income'])
    expense_count = len([op for op in operations if op['type'] == 'expense'])
    
    income_formatted = f"{income_total:,.0f} â‚½".replace(',', ' ')
    expense_formatted = f"{expense_total:,.0f} â‚½".replace(',', ' ')
    profit_formatted = f"{income_total - expense_total:,.0f} â‚½".replace(',', ' ')
    
    ml6_total = sum(op['amount'] for op in operations if op['type'] == 'income' and op.get('volume') == '6ml')
    ml10_total = sum(op['amount'] for op in operations if op['type'] == 'income' and op.get('volume') == '10ml')
    
    parfums = {}
    for op in operations:
        if op['type'] == 'income':
            key = op['parfum_name']
            if key not in parfums:
                parfums[key] = {
                    'amount': 0,
                    'quantity': 0
                }
            parfums[key]['amount'] += op['amount']
            parfums[key]['quantity'] += op['quantity']
    
    report = "ğŸ“Š *ĞĞ‘Ğ©ĞĞ¯ Ğ¡Ğ¢ĞĞ¢Ğ˜Ğ¡Ğ¢Ğ˜ĞšĞ*\n"
    report += "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
    report += f"ğŸ“ˆ *Ğ”Ğ¾Ñ…Ğ¾Ğ´Ñ‹:* `{income_formatted}` ({income_count} ÑˆÑ‚)\n"
    report += f"ğŸ“‰ *Ğ Ğ°ÑÑ…Ğ¾Ğ´Ñ‹:* `{expense_formatted}` ({expense_count} ÑˆÑ‚)\n"
    report += f"ğŸ’ *Ğ˜Ñ‚Ğ¾Ğ³:* `{profit_formatted}`\n\n"
    report += f"ğŸ“¦ *ĞŸĞ¾ Ğ¾Ğ±ÑŠĞµĞ¼Ñƒ:*\n"
    report += f"   â€¢ 6ml: {ml6_total:,.0f} â‚½\n".replace(',', ' ')
    report += f"   â€¢ 10ml: {ml10_total:,.0f} â‚½\n".replace(',', ' ')
    
    if parfums:
        report += f"\nğŸ† *Ğ¢Ğ¾Ğ¿ Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ¾Ğ²:*\n"
        top_parfums = sorted(parfums.items(), key=lambda x: x[1]['amount'], reverse=True)[:5]
        for parfum, data in top_parfums:
            pamount = f"{data['amount']:,.0f} â‚½".replace(',', ' ')
            report += f"   â€¢ {parfum}: {data['quantity']} ÑˆÑ‚ ({pamount})\n"
    
    update.message.reply_text(report, parse_mode='Markdown', reply_markup=main_keyboard)

def show_operations_for_edit(update, context):
    if not check_access(update):
        return
    
    if update.message.text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        update.message.reply_text("ğŸ”™ Ğ’Ğ¾Ğ·Ğ²Ñ€Ğ°Ñ‚ Ğ² Ğ³Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    operations = get_all_operations()
    
    if not operations:
        update.message.reply_text("ğŸ“­ ĞĞµÑ‚ Ğ¾Ğ¿ĞµÑ€Ğ°Ñ†Ğ¸Ğ¹", reply_markup=main_keyboard)
        return
    
    operations.sort(key=lambda x: x['id'], reverse=True)
    operations = operations[:15]
    
    keyboard = []
    for op in operations:
        amount = op['amount']
        formatted_amount = f"{amount:,.0f} â‚½".replace(',', ' ')
        if amount != int(amount):
            formatted_amount = f"{amount:,.2f} â‚½".replace(',', ' ')
        
        if op['type'] == 'income':
            desc = f"{op['parfum_name']} {op['volume']} x{op['quantity']} - {op['employee']}"
        else:
            desc = op['description'][:20] + "..." if len(op['description']) > 20 else op['description']
        
        button_text = f"#{op['id']} {op['type_display']} {formatted_amount} - {desc}"
        keyboard.append([InlineKeyboardButton(button_text, callback_data=f"edit_op_{op['id']}")])
    
    keyboard.append([InlineKeyboardButton("ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°", callback_data="edit_cancel")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    update.message.reply_text(
        "âœï¸ *Ğ’Ğ«Ğ‘Ğ•Ğ Ğ˜Ğ¢Ğ• ĞĞŸĞ•Ğ ĞĞ¦Ğ˜Ğ®:*\n"
        "â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
        "ğŸ”¹ *ĞŸĞ¾ÑĞ»ĞµĞ´Ğ½Ğ¸Ğµ 15 Ğ¾Ğ¿ĞµÑ€Ğ°Ñ†Ğ¸Ğ¹:*",
        parse_mode='Markdown',
        reply_markup=reply_markup
    )

def edit_callback(update, context):
    query = update.callback_query
    query.answer()
    
    if not check_access(update):
        query.edit_message_text("âŒ ĞĞµÑ‚ Ğ´Ğ¾ÑÑ‚ÑƒĞ¿Ğ°")
        return
    
    data = query.data
    
    if data == "edit_cancel":
        query.edit_message_text("ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½ĞµĞ½Ğ¾")
        return
    
    if data.startswith("edit_op_"):
        op_id = int(data.split('_')[2])
        
        operations = get_all_operations()
        op = next((o for o in operations if o['id'] == op_id), None)
        
        if not op:
            query.edit_message_text("âŒ ĞĞ¿ĞµÑ€Ğ°Ñ†Ğ¸Ñ Ğ½Ğµ Ğ½Ğ°Ğ¹Ğ´ĞµĞ½Ğ°")
            return
        
        amount = op['amount']
        formatted_amount = f"{amount:,.0f} â‚½".replace(',', ' ')
        if amount != int(amount):
            formatted_amount = f"{amount:,.2f} â‚½".replace(',', ' ')
        
        if op['type'] == 'income':
            op_text = (
                f"ğŸ“Œ *ĞŸĞ ĞĞ”ĞĞ–Ğ #{op_id}*\n"
                f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
                f"ğŸ“… Ğ”Ğ°Ñ‚Ğ°: {op['date']}\n"
                f"ğŸ“¦ ĞŸĞ°Ñ€Ñ„ÑĞ¼: {op['parfum_name']}\n"
                f"ğŸ”¢ ĞĞ±ÑŠĞµĞ¼: {op['volume']}\n"
                f"ğŸ“Š ĞšĞ¾Ğ»-Ğ²Ğ¾: {op['quantity']} ÑˆÑ‚\n"
                f"ğŸ‘¤ Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº: {op['employee']}\n"
                f"ğŸ’³ ĞĞ¿Ğ»Ğ°Ñ‚Ğ°: {op['payment']}\n"
                f"ğŸ¦ Ğ‘Ğ°Ğ½Ğº: {op.get('bank', '-')}\n"
                f"ğŸ’° Ğ¡ÑƒĞ¼Ğ¼Ğ°: {formatted_amount}\n"
                f"ğŸ“ Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ğ»: {op['added_by']}"
            )
        else:
            op_text = (
                f"ğŸ“Œ *Ğ ĞĞ¡Ğ¥ĞĞ” #{op_id}*\n"
                f"â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n\n"
                f"ğŸ“… Ğ”Ğ°Ñ‚Ğ°: {op['date']}\n"
                f"ğŸ’° Ğ¡ÑƒĞ¼Ğ¼Ğ°: {formatted_amount}\n"
                f"ğŸ“‹ ĞĞ¿Ğ¸ÑĞ°Ğ½Ğ¸Ğµ: {op['description']}\n"
                f"ğŸ‘¤ Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº: {op['employee']}\n"
                f"ğŸ“ Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ¸Ğ»: {op['added_by']}"
            )
        
        keyboard = [
            [InlineKeyboardButton("ğŸ’° Ğ˜Ğ·Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ ÑÑƒĞ¼Ğ¼Ñƒ", callback_data=f"edit_sum_{op_id}")],
            [InlineKeyboardButton("âŒ Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ Ğ¾Ğ¿ĞµÑ€Ğ°Ñ†Ğ¸Ñ", callback_data=f"edit_del_{op_id}")]
        ]
        
        if op['type'] == 'income':
            keyboard.insert(1, [InlineKeyboardButton("ğŸ‘¤ Ğ˜Ğ·Ğ¼ĞµĞ½Ğ¸Ñ‚ÑŒ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°", callback_data=f"edit_employee_{op_id}")])
        
        keyboard.append([InlineKeyboardButton("ğŸ”™ ĞĞ°Ğ·Ğ°Ğ´", callback_data="edit_back")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        query.edit_message_text(op_text, parse_mode='Markdown', reply_markup=reply_markup)
    
    elif data.startswith("edit_sum_"):
        op_id = int(data.split('_')[2])
        context.user_data['edit_op_id'] = op_id
        context.user_data['edit_action'] = 'sum'
        query.edit_message_text(f"âœï¸ Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ğ½Ğ¾Ğ²ÑƒÑ ÑÑƒĞ¼Ğ¼Ñƒ Ğ´Ğ»Ñ Ğ¾Ğ¿ĞµÑ€Ğ°Ñ†Ğ¸Ğ¸ #{op_id}:")
    
    elif data.startswith("edit_employee_"):
        op_id = int(data.split('_')[2])
        
        keyboard = []
        for emp in EMPLOYEES:
            keyboard.append([InlineKeyboardButton(f"ğŸ‘¤ {emp}", callback_data=f"edit_set_employee_{op_id}_{emp}")])
        keyboard.append([InlineKeyboardButton("ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°", callback_data=f"edit_op_{op_id}")])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        query.edit_message_text(f"âœï¸ Ğ’Ñ‹Ğ±ĞµÑ€Ğ¸Ñ‚Ğµ Ğ½Ğ¾Ğ²Ğ¾Ğ³Ğ¾ ÑĞ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸ĞºĞ°:", reply_markup=reply_markup)
    
    elif data.startswith("edit_set_employee_"):
        parts = data.split('_')
        op_id = int(parts[3])
        new_employee = parts[4]
        
        operations = get_all_operations()
        for i, op in enumerate(operations):
            if op['id'] == op_id:
                op['employee'] = new_employee
                update_operation(op_id, op)
                break
        
        query.edit_message_text(f"âœ… Ğ¡Ğ¾Ñ‚Ñ€ÑƒĞ´Ğ½Ğ¸Ğº Ğ¸Ğ·Ğ¼ĞµĞ½ĞµĞ½ Ğ½Ğ° {new_employee}")
    
    elif data.startswith("edit_del_"):
        op_id = int(data.split('_')[2])
        
        keyboard = [
            [InlineKeyboardButton("âœ… Ğ”Ğ°", callback_data=f"edit_confirm_del_{op_id}")],
            [InlineKeyboardButton("âŒ ĞĞµÑ‚", callback_data=f"edit_op_{op_id}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        query.edit_message_text(f"âš ï¸ Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ Ğ¾Ğ¿ĞµÑ€Ğ°Ñ†Ğ¸Ñ #{op_id}?", reply_markup=reply_markup)
    
    elif data.startswith("edit_confirm_del_"):
        op_id = int(data.split('_')[3])
        delete_operation(op_id)
        query.edit_message_text(f"âœ… ĞĞ¿ĞµÑ€Ğ°Ñ†Ğ¸Ñ #{op_id} ÑƒĞ´Ğ°Ğ»ĞµĞ½Ğ°")
    
    elif data == "edit_back":
        show_operations_for_edit(update, context)

def handle_edit_input(update, context):
    if 'edit_op_id' not in context.user_data:
        return
    
    if update.message.text == '/cancel':
        del context.user_data['edit_op_id']
        del context.user_data['edit_action']
        update.message.reply_text("ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½ĞµĞ½Ğ¾", reply_markup=main_keyboard)
        return
    
    op_id = context.user_data['edit_op_id']
    action = context.user_data['edit_action']
    text = update.message.text
    
    if action == 'sum':
        try:
            new_sum = float(text.replace(' ', '').replace(',', '.'))
            
            operations = get_all_operations()
            for op in operations:
                if op['id'] == op_id:
                    op['amount'] = new_sum
                    update_operation(op_id, op)
                    break
            
            formatted_sum = f"{new_sum:,.0f} â‚½".replace(',', ' ')
            if new_sum != int(new_sum):
                formatted_sum = f"{new_sum:,.2f} â‚½".replace(',', ' ')
            
            update.message.reply_text(f"âœ… Ğ¡ÑƒĞ¼Ğ¼Ğ° Ğ¸Ğ·Ğ¼ĞµĞ½ĞµĞ½Ğ° Ğ½Ğ° {formatted_sum}", reply_markup=main_keyboard)
            
        except ValueError:
            update.message.reply_text("âŒ Ğ’Ğ²ĞµĞ´Ğ¸Ñ‚Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾")
            return
    
    del context.user_data['edit_op_id']
    del context.user_data['edit_action']

def handle_message(update, context):
    if not check_access(update):
        update.message.reply_text("âŒ ĞĞµÑ‚ Ğ´Ğ¾ÑÑ‚ÑƒĞ¿Ğ°")
        return
    
    chat_id = update.effective_chat.id
    text = update.message.text
    
    if text == 'ğŸ”™ ĞÑ‚Ğ¼ĞµĞ½Ğ°':
        if chat_id in user_data:
            del user_data[chat_id]
        context.user_data.clear()
        update.message.reply_text("ğŸ”™ Ğ“Ğ»Ğ°Ğ²Ğ½Ğ¾Ğµ Ğ¼ĞµĞ½Ñ", reply_markup=main_keyboard)
        return
    
    if 'edit_op_id' in context.user_data:
        handle_edit_input(update, context)
        return
    
    if chat_id in user_data:
        state_data = user_data[chat_id]
        
        if state_data.get('type') == 'income':
            state = state_data.get('state')
            
            if state == INCOME_STATES['NAME']:
                handle_income_name(update, context)
            elif state == INCOME_STATES['VOLUME']:
                handle_income_volume(update, context)
            elif state == INCOME_STATES['QUANTITY']:
                handle_income_quantity(update, context)
            elif state == INCOME_STATES['EMPLOYEE']:
                handle_income_employee(update, context)
            elif state == INCOME_STATES['PAYMENT']:
                handle_income_payment(update, context)
            elif state == INCOME_STATES['BANK']:
                handle_income_bank(update, context)
            elif state == INCOME_STATES['AMOUNT']:
                handle_income_amount(update, context)
        
        elif state_data.get('type') == 'expense':
            state = state_data.get('state')
            
            if state == EXPENSE_STATES['AMOUNT']:
                handle_expense_amount(update, context)
            elif state == EXPENSE_STATES['DESCRIPTION']:
                handle_expense_description(update, context)
            elif state == EXPENSE_STATES['EMPLOYEE']:
                handle_expense_employee(update, context)
    
    elif text == 'ğŸ’° Ğ”Ğ¾Ñ…Ğ¾Ğ´':
        handle_income(update, context)
    elif text == 'ğŸ’¸ Ğ Ğ°ÑÑ…Ğ¾Ğ´':
        handle_expense(update, context)
    elif text == 'ğŸ“Š Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ°':
        show_all_statistics(update, context)
    elif text == 'ğŸ“‹ Ğ¢Ğ°Ğ±Ğ»Ğ¸Ñ†Ğ° Ğ¿Ğ°Ñ€Ñ„ÑĞ¼Ğ¾Ğ²':
        show_parfum_table(update, context)
    elif text == 'ğŸ‘¥ Ğ¡Ñ‚Ğ°Ñ‚Ğ¸ÑÑ‚Ğ¸ĞºĞ° ĞºĞ¾Ğ»Ğ»ĞµĞ³':
        show_employee_stats(update, context)
    elif text == 'âœï¸ Ğ ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ñ‚ÑŒ/Ğ£Ğ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ':
        show_operations_for_edit(update, context)
    else:
        update.message.reply_text("Ğ˜ÑĞ¿Ğ¾Ğ»ÑŒĞ·ÑƒĞ¹Ñ‚Ğµ ĞºĞ½Ğ¾Ğ¿ĞºĞ¸ ğŸ‘‡", reply_markup=main_keyboard)

def main():
    print("âœ… Ğ‘Ğ¾Ñ‚ Ğ·Ğ°Ğ¿ÑƒÑĞºĞ°ĞµÑ‚ÑÑ...")
    init_excel()
    print("âœ… Ğ”Ğ°Ğ½Ğ½Ñ‹Ğµ Ğ±ÑƒĞ´ÑƒÑ‚ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ÑÑ‚ÑŒÑÑ Ğ² gabbana_data.json Ğ¸ gabbana_budget.xlsx")
    
    updater = Updater(TOKEN, use_context=True)
    dp = updater.dispatcher
    
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(CallbackQueryHandler(edit_callback, pattern="^edit_"))
    dp.add_handler(MessageHandler(Filters.text, handle_message))
    
    print("âœ… Ğ‘Ğ¾Ñ‚ Ğ³Ğ¾Ñ‚Ğ¾Ğ² Ğº Ñ€Ğ°Ğ±Ğ¾Ñ‚Ğµ!")
    
    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\nâŒ ĞÑˆĞ¸Ğ±ĞºĞ°: {e}")
